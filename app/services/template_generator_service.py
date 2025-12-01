"""
Сервис генерации Excel шаблонов для поставщиков
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO
from pathlib import Path
from app import db
from app.models.supplier import Supplier
from app.models.subcategory import Subcategory
from app.models.attribute import AttributeType


class TemplateGeneratorService:
    """Сервис для генерации Excel шаблонов для заполнения поставщиками"""
    
    @staticmethod
    def generate_supplier_template(supplier_id, category_id=None):
        """
        Генерирует Excel шаблон для поставщика
        
        Args:
            supplier_id: ID поставщика
            category_id: ID категории (опционально, если None - все категории поставщика)
        
        Returns:
            BytesIO: Поток с Excel файлом
        """
        supplier = Supplier.query.get_or_404(supplier_id)
        
        # Получить подкатегории поставщика
        if category_id:
            # Только подкатегории указанной категории
            subcategories = Subcategory.query.join(
                'suppliers'
            ).filter(
                Supplier.id == supplier_id,
                Subcategory.category_id == category_id,
                Subcategory.is_active == True
            ).order_by(Subcategory.code).all()
        else:
            # Все подкатегории поставщика
            subcategories = supplier.subcategories.filter(
                Subcategory.is_active == True
            ).order_by(Subcategory.code).all()
        
        if not subcategories:
            raise ValueError("У поставщика нет активных подкатегорий")
        
        # Создать Excel файл
        wb = Workbook()
        wb.remove(wb.active)  # Удалить дефолтный лист
        
        # Создать лист с инструкциями
        instructions_sheet = wb.create_sheet("📋 ИНСТРУКЦИЯ", 0)
        TemplateGeneratorService._add_instructions(instructions_sheet, supplier, subcategories)
        
        # Создать лист для каждой подкатегории
        for subcategory in subcategories:
            sheet = wb.create_sheet(subcategory.name[:31])  # Excel ограничение 31 символ
            TemplateGeneratorService._add_subcategory_sheet(sheet, subcategory)
        
        # Сохранить в BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def _add_instructions(sheet, supplier, subcategories):
        """Добавить лист с инструкциями"""
        # Заголовок
        sheet['A1'] = f"ШАБЛОН ДЛЯ ЗАПОЛНЕНИЯ - {supplier.name}"
        sheet['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        sheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells('A1:D1')
        
        row = 3
        
        # Общая информация
        sheet[f'A{row}'] = "ОБЩАЯ ИНФОРМАЦИЯ"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        sheet[f'A{row}'] = "Поставщик:"
        sheet[f'B{row}'] = supplier.name
        row += 1
        
        sheet[f'A{row}'] = "Код поставщика:"
        sheet[f'B{row}'] = supplier.code
        row += 2
        
        # Инструкции
        sheet[f'A{row}'] = "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        instructions = [
            "1. В файле есть отдельный лист для каждой подкатегории товаров",
            "2. На каждом листе заполните информацию о товарах:",
            "   - Каждая строка = один товар",
            "   - Обязательные поля помечены красным фоном",
            "   - Поля с единицами измерения: укажите только число (единица уже указана)",
            "",
            "3. Для фото и 3D моделей:",
            "   - Укажите URL (ссылку) на файл в интернете",
            "   - Можно указать несколько URL через запятую",
            "   - Пример: https://example.com/photo1.jpg, https://example.com/photo2.jpg",
            "",
            "4. Для атрибутов типа 'Выбор' (SELECT):",
            "   - Используйте выпадающий список в ячейке",
            "   - Выберите значение из предложенных вариантов",
            "",
            "5. После заполнения:",
            "   - Сохраните файл",
            "   - Отправьте файл обратно для импорта",
            "",
            "ВАЖНО:",
            "- Не удаляйте и не переименовывайте листы",
            "- Не изменяйте названия колонок (заголовки)",
            "- Не добавляйте новые колонки",
            "- Обязательные поля должны быть заполнены",
        ]
        
        for instruction in instructions:
            sheet[f'A{row}'] = instruction
            sheet[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
            if instruction.startswith("ВАЖНО:"):
                sheet[f'A{row}'].font = Font(bold=True, color="FF0000")
            row += 1
        
        row += 1
        
        # Список подкатегорий
        sheet[f'A{row}'] = "СПИСОК ПОДКАТЕГОРИЙ В ШАБЛОНЕ"
        sheet[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        sheet[f'A{row}'] = "№"
        sheet[f'B{row}'] = "Код"
        sheet[f'C{row}'] = "Название"
        sheet[f'D{row}'] = "Атрибутов"
        for col in ['A', 'B', 'C', 'D']:
            sheet[f'{col}{row}'].font = Font(bold=True)
            sheet[f'{col}{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        row += 1
        
        for idx, subcat in enumerate(subcategories, 1):
            sheet[f'A{row}'] = idx
            sheet[f'B{row}'] = subcat.code
            sheet[f'C{row}'] = subcat.name
            sheet[f'D{row}'] = subcat.attributes.count()
            row += 1
        
        # Настроить ширину колонок
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 15
    
    @staticmethod
    def _add_subcategory_sheet(sheet, subcategory):
        """Добавить лист для подкатегории"""
        # Заголовок листа
        sheet['A1'] = f"{subcategory.code} - {subcategory.name}"
        sheet['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        sheet['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Получить все атрибуты подкатегории
        subcat_attrs = subcategory.get_all_attributes()
        
        # Определить колонки
        columns = []
        
        # Базовые обязательные поля
        columns.append({
            'name': 'Артикул производителя (SKU)',
            'code': 'sku',
            'required': True,
            'type': AttributeType.TEXT,
            'unit': None,
            'description': 'Уникальный артикул товара от производителя'
        })
        
        columns.append({
            'name': 'Название товара',
            'code': 'name',
            'required': True,
            'type': AttributeType.TEXT,
            'unit': None,
            'description': 'Полное название товара'
        })
        
        columns.append({
            'name': 'Описание',
            'code': 'description',
            'required': False,
            'type': AttributeType.TEXT,
            'unit': None,
            'description': 'Подробное описание товара'
        })
        
        # Добавить атрибуты подкатегории
        for subcat_attr in subcat_attrs:
            attr = subcat_attr.attribute
            columns.append({
                'name': attr.name + (f" ({attr.unit})" if attr.unit else ""),
                'code': attr.code,
                'required': subcat_attr.is_required,
                'type': attr.type,
                'unit': attr.unit,
                'description': attr.description or "",
                'select_values': [v.value for v in attr.values.all()] if attr.type == AttributeType.SELECT else None
            })
        
        # Записать заголовки
        header_row = 3
        for col_idx, col_info in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            cell = sheet[f'{col_letter}{header_row}']
            cell.value = col_info['name']
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Красный фон для обязательных полей
            if col_info['required']:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="9C0006")
            else:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Добавить комментарий с описанием
            if col_info.get('description'):
                cell.comment = col_info['description']
            
            # Настроить ширину колонки
            sheet.column_dimensions[col_letter].width = max(15, len(col_info['name']) + 2)
        
        # Добавить пример строки с подсказками
        example_row = header_row + 1
        sheet[f'A{example_row}'] = "ПРИМЕР (удалите эту строку перед заполнением)"
        sheet[f'A{example_row}'].font = Font(italic=True, color="808080")
        sheet[f'A{example_row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        example_row += 1
        for col_idx, col_info in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            cell = sheet[f'{col_letter}{example_row}']
            
            # Примеры значений в зависимости от типа
            if col_info['code'] == 'sku':
                cell.value = "SKU-001"
            elif col_info['code'] == 'name':
                cell.value = "Пример товара"
            elif col_info['code'] == 'description':
                cell.value = "Подробное описание товара..."
            elif col_info['type'] == AttributeType.TEXT:
                cell.value = "Текст"
            elif col_info['type'] == AttributeType.NUMBER:
                cell.value = "100" + (f" {col_info['unit']}" if col_info['unit'] else "")
            elif col_info['type'] == AttributeType.DATE:
                cell.value = "2024-01-01"
            elif col_info['type'] == AttributeType.BOOLEAN:
                cell.value = "Да / Нет"
            elif col_info['type'] == AttributeType.URL:
                # Может быть 3D модель или другой URL
                if '3d' in col_info['code'].lower() or 'модель' in col_info['name'].lower():
                    cell.value = "https://example.com/model.glb"
                else:
                    cell.value = "https://example.com"
            elif col_info['type'] == AttributeType.IMAGE:
                cell.value = "https://example.com/photo.jpg"
            elif col_info['type'] == AttributeType.SELECT:
                if col_info.get('select_values'):
                    cell.value = col_info['select_values'][0] if col_info['select_values'] else ""
                else:
                    cell.value = "Выберите из списка"
            
            cell.font = Font(italic=True, color="808080")
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Добавить выпадающие списки для SELECT атрибутов
        for col_idx, col_info in enumerate(columns, start=1):
            if col_info['type'] == AttributeType.SELECT and col_info.get('select_values'):
                col_letter = get_column_letter(col_idx)
                # Создать валидацию данных
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(col_info["select_values"])}"',
                    allow_blank=not col_info['required']
                )
                dv.error = "Выберите значение из списка"
                dv.errorTitle = "Неверное значение"
                sheet.add_data_validation(dv)
                # Применить к колонке (начиная со строки после примера)
                dv.add(f'{col_letter}{example_row + 1}:{col_letter}1000')
        
        # Заморозить заголовки
        sheet.freeze_panes = f'A{example_row + 1}'
        
        # Добавить границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(header_row, example_row + 1):
            for col_idx in range(1, len(columns) + 1):
                col_letter = get_column_letter(col_idx)
                sheet[f'{col_letter}{row}'].border = thin_border

